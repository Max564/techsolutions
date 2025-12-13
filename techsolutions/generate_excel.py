#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

def create_product_catalog():
    """Crée le catalogue produit complet"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Catalogue Produits"
    
    # En-têtes
    headers = ['Référence', 'Catégorie', 'Désignation', 'Prix HT (€)', 'Fournisseur']
    ws.append(headers)
    
    # Style des en-têtes
    header_fill = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Données du catalogue
    products = [
        # COMPOSANTS PC
        ["CPU001", "Processeur", "Intel Core i5-13400F (10 cœurs, 4.6GHz)", 219.99, "LDLC"],
        ["CPU002", "Processeur", "Intel Core i7-13700K (16 cœurs, 5.4GHz)", 419.99, "LDLC"],
        ["CPU003", "Processeur", "AMD Ryzen 5 7600X (6 cœurs, 5.3GHz)", 249.99, "TopAchat"],
        ["CPU004", "Processeur", "AMD Ryzen 7 7800X3D (8 cœurs, 5.0GHz)", 449.99, "TopAchat"],
        
        ["CM001", "Carte Mère", "MSI B660M PRO-VDH WiFi (Socket LGA1700)", 129.99, "LDLC"],
        ["CM002", "Carte Mère", "MSI Z790 TOMAHAWK WiFi (Socket LGA1700)", 299.99, "LDLC"],
        ["CM003", "Carte Mère", "ASUS TUF Gaming B650-PLUS WiFi (AM5)", 199.99, "TopAchat"],
        ["CM004", "Carte Mère", "GIGABYTE X670 AORUS Elite AX (AM5)", 289.99, "TopAchat"],
        
        ["RAM001", "Mémoire RAM", "Corsair Vengeance 16Go (2x8Go) DDR4 3200MHz", 54.99, "LDLC"],
        ["RAM002", "Mémoire RAM", "Corsair Vengeance 32Go (2x16Go) DDR4 3600MHz", 89.99, "LDLC"],
        ["RAM003", "Mémoire RAM", "G.Skill Trident Z5 32Go (2x16Go) DDR5 6000MHz", 149.99, "TopAchat"],
        ["RAM004", "Mémoire RAM", "G.Skill Trident Z5 64Go (2x32Go) DDR5 6400MHz", 299.99, "TopAchat"],
        
        ["SSD001", "Stockage SSD", "Samsung 970 EVO Plus 500Go M.2 NVMe", 59.99, "LDLC"],
        ["SSD002", "Stockage SSD", "Samsung 980 PRO 1To M.2 NVMe PCIe 4.0", 119.99, "LDLC"],
        ["SSD003", "Stockage SSD", "WD Black SN850X 2To M.2 NVMe PCIe 4.0", 189.99, "TopAchat"],
        ["SSD004", "Stockage SSD", "Crucial P5 Plus 4To M.2 NVMe PCIe 4.0", 349.99, "TopAchat"],
        
        ["HDD001", "Stockage HDD", "Seagate BarraCuda 2To 7200 RPM SATA", 54.99, "LDLC"],
        ["HDD002", "Stockage HDD", "WD Blue 4To 5400 RPM SATA", 89.99, "TopAchat"],
        
        ["GPU001", "Carte Graphique", "NVIDIA RTX 4060 8Go GDDR6", 329.99, "LDLC"],
        ["GPU002", "Carte Graphique", "NVIDIA RTX 4070 12Go GDDR6X", 629.99, "LDLC"],
        ["GPU003", "Carte Graphique", "AMD Radeon RX 7700 XT 12Go", 449.99, "TopAchat"],
        ["GPU004", "Carte Graphique", "NVIDIA RTX 4080 16Go GDDR6X", 1199.99, "TopAchat"],
        
        ["PSU001", "Alimentation", "Corsair RM650x 650W 80+ Gold Modulaire", 99.99, "LDLC"],
        ["PSU002", "Alimentation", "be quiet! Straight Power 11 750W 80+ Gold", 129.99, "LDLC"],
        ["PSU003", "Alimentation", "Seasonic PRIME TX-850 850W 80+ Titanium", 189.99, "TopAchat"],
        
        ["BOI001", "Boîtier", "Fractal Design Define 7 Compact (Noir)", 119.99, "LDLC"],
        ["BOI002", "Boîtier", "NZXT H510 Flow (Blanc)", 99.99, "LDLC"],
        ["BOI003", "Boîtier", "Corsair 4000D Airflow (Noir)", 104.99, "TopAchat"],
        ["BOI004", "Boîtier", "be quiet! Pure Base 500DX (Blanc)", 109.99, "TopAchat"],
        
        ["VEN001", "Refroidissement", "Noctua NH-U12S Redux Ventirad CPU", 49.99, "LDLC"],
        ["VEN002", "Refroidissement", "be quiet! Dark Rock Pro 4 Ventirad CPU", 84.99, "LDLC"],
        ["VEN003", "Refroidissement", "Corsair iCUE H100i RGB Elite Watercooling 240mm", 134.99, "TopAchat"],
        
        # PÉRIPHÉRIQUES
        ["ECR001", "Écran", "Dell P2422H 24\" Full HD IPS 60Hz", 179.99, "LDLC"],
        ["ECR002", "Écran", "LG 27UK850-W 27\" 4K IPS HDR", 449.99, "LDLC"],
        ["ECR003", "Écran", "ASUS ProArt PA278QV 27\" WQHD IPS", 349.99, "TopAchat"],
        ["ECR004", "Écran", "BenQ SW271C 27\" 4K IPS Adobe RGB", 899.99, "TopAchat"],
        
        ["ECRA001", "Écran Accessibilité", "Samsung ViewFinity S8 32\" 4K avec lecteur d'écran", 699.99, "LDLC"],
        ["ECRA002", "Écran Accessibilité", "Dell UltraSharp 27\" avec grossissement intégré", 549.99, "LDLC"],
        
        ["CLV001", "Clavier", "Logitech K120 Clavier filaire", 14.99, "LDLC"],
        ["CLV002", "Clavier", "Microsoft Sculpt Ergonomic Desktop", 89.99, "LDLC"],
        ["CLV003", "Clavier", "Logitech MX Keys Clavier sans fil rétroéclairé", 109.99, "TopAchat"],
        ["CLV004", "Clavier", "Keychron K8 Pro Mécanique sans fil", 119.99, "TopAchat"],
        
        ["CLVA001", "Clavier Accessibilité", "Clavier grands caractères contrasté avec rétroéclairage", 89.99, "Handytech"],
        ["CLVA002", "Clavier Accessibilité", "Clavier Braille 40 cellules refresh", 899.99, "Handytech"],
        
        ["SOU001", "Souris", "Logitech M185 Souris sans fil", 12.99, "LDLC"],
        ["SOU002", "Souris", "Logitech MX Master 3S Souris sans fil", 109.99, "LDLC"],
        ["SOU003", "Souris", "Microsoft Ergonomic Mouse", 44.99, "TopAchat"],
        
        ["SOUA001", "Souris Accessibilité", "Souris trackball ergonomique Kensington Expert", 119.99, "Handytech"],
        ["SOUA002", "Souris Accessibilité", "Contacteur bouton adapté grand format", 79.99, "Handytech"],
        
        ["IMP001", "Imprimante", "HP LaserJet Pro M404dn Mono", 229.99, "LDLC"],
        ["IMP002", "Imprimante", "Brother MFC-L3770CDW Couleur Multifonction", 399.99, "LDLC"],
        ["IMP003", "Imprimante", "Epson EcoTank ET-2820 Jet d'encre Couleur", 249.99, "TopAchat"],
        
        ["WEB001", "Webcam", "Logitech C920 HD Pro 1080p", 74.99, "LDLC"],
        ["WEB002", "Webcam", "Logitech Brio 4K Ultra HD", 199.99, "TopAchat"],
        
        ["CAS001", "Casque/Micro", "Logitech H390 USB", 34.99, "LDLC"],
        ["CAS002", "Casque/Micro", "Jabra Evolve2 65 USB-A/Bluetooth", 229.99, "LDLC"],
        
        ["CASA001", "Casque Accessibilité", "Casque audio amplifié pour malentendants", 159.99, "Handytech"],
        
        # LOGICIELS
        ["OS001", "Système", "Windows 11 Pro (OEM)", 149.99, "Microsoft"],
        ["OS002", "Système", "Windows 11 Pro (Licence Volume)", 199.99, "Microsoft"],
        
        ["OFF001", "Suite Bureautique", "Microsoft 365 Business Standard (1 an)", 129.99, "Microsoft"],
        ["OFF002", "Suite Bureautique", "Microsoft Office 2021 Famille et Petite Entreprise", 299.99, "Microsoft"],
        
        ["OFFL001", "Suite Bureautique", "LibreOffice Suite complète (Gratuit)", 0.00, "LibreOffice"],
        
        ["AV001", "Antivirus", "Bitdefender Total Security (3 appareils, 1 an)", 49.99, "Bitdefender"],
        ["AV002", "Antivirus", "Kaspersky Total Security (5 appareils, 1 an)", 54.99, "Kaspersky"],
        ["AV003", "Antivirus", "Norton 360 Deluxe (5 appareils, 1 an)", 44.99, "Norton"],
        
        ["AVL001", "Antivirus", "Avast Free Antivirus (Gratuit)", 0.00, "Avast"],
        
        ["LOG001", "Logiciel Pro", "Adobe Creative Cloud All Apps (1 an)", 719.88, "Adobe"],
        ["LOG002", "Logiciel Pro", "Autodesk AutoCAD LT (1 an)", 499.00, "Autodesk"],
        ["LOG003", "Logiciel Pro", "Slack Business+ (par utilisateur/an)", 149.00, "Slack"],
        
        ["LOGA001", "Logiciel Accessibilité", "JAWS Lecteur d'écran (Licence)", 1095.00, "Freedom Scientific"],
        ["LOGA002", "Logiciel Accessibilité", "ZoomText Magnifier Grossisseur (Licence)", 599.00, "Freedom Scientific"],
        ["LOGA003", "Logiciel Accessibilité", "NVDA Lecteur d'écran (Gratuit Open Source)", 0.00, "NV Access"],
        
        # RÉSEAU
        ["ROU001", "Routeur", "TP-Link Archer AX53 WiFi 6 (AX3000)", 79.99, "LDLC"],
        ["ROU002", "Routeur", "ASUS RT-AX86U Pro WiFi 6E", 249.99, "LDLC"],
        ["ROU003", "Routeur", "Netgear Nighthawk RAX200 WiFi 6 (AX11000)", 449.99, "TopAchat"],
        
        ["SWI001", "Switch", "TP-Link TL-SG108 Switch 8 ports Gigabit", 24.99, "LDLC"],
        ["SWI002", "Switch", "Netgear GS316 Switch 16 ports Gigabit", 89.99, "LDLC"],
        ["SWI003", "Switch", "Cisco SG350-28P Switch 28 ports PoE+", 599.99, "Cisco"],
        
        ["CAB001", "Câbles", "Lot 10 câbles RJ45 Cat6 1m", 19.99, "LDLC"],
        ["CAB002", "Câbles", "Lot 5 câbles RJ45 Cat6A 3m", 29.99, "LDLC"],
        
        ["FW001", "Firewall", "Fortinet FortiGate 60F Firewall", 1299.99, "Fortinet"],
        ["FW002", "Firewall", "pfSense SG-1100 Appliance", 199.99, "Netgate"],
        
        # SÉCURITÉ
        ["VPN001", "VPN", "NordVPN Business (par utilisateur/an)", 89.00, "NordVPN"],
        ["VPN002", "VPN", "ExpressVPN Business (par utilisateur/an)", 119.95, "ExpressVPN"],
        
        ["BAC001", "Sauvegarde", "Synology DS220+ NAS 2 baies", 329.99, "Synology"],
        ["BAC002", "Sauvegarde", "QNAP TS-464 NAS 4 baies", 599.99, "QNAP"],
        ["BAC003", "Sauvegarde", "Acronis Cyber Protect Standard (1 poste/an)", 79.99, "Acronis"],
        
        # SERVICES
        ["SER001", "Service", "Installation et configuration poste de travail (par poste)", 120.00, "TechSolutions"],
        ["SER002", "Service", "Migration de données et applications", 200.00, "TechSolutions"],
        ["SER003", "Service", "Formation utilisateur (par journée)", 600.00, "TechSolutions"],
        ["SER004", "Service", "Support technique à distance (par heure)", 80.00, "TechSolutions"],
        ["SER005", "Service", "Audit de sécurité complet", 1500.00, "TechSolutions"],
        ["SER006", "Service", "Configuration réseau entreprise", 800.00, "TechSolutions"],
        ["SER007", "Service", "Garantie extension 3 ans (par poste)", 150.00, "TechSolutions"],
    ]
    
    for row in products:
        ws.append(row)
    
    # Ajuster les largeurs de colonnes
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 60
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 20
    
    # Format prix
    for row in ws.iter_rows(min_row=2, min_col=4, max_col=4):
        for cell in row:
            cell.number_format = '#,##0.00 €'
    
    wb.save('/home/claude/techsolutions/Catalogue_produit_TechSolutions.xlsx')
    print("✓ Catalogue produit créé avec succès")
    return wb

def create_automated_quote_template():
    """Crée le modèle de devis automatique"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Devis"
    
    # Logo et en-tête
    ws.merge_cells('A1:F1')
    ws['A1'] = 'TECHSOLUTIONS'
    ws['A1'].font = Font(size=24, bold=True, color="4A90E2")
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 40
    
    # Informations entreprise
    ws['A3'] = 'TechSolutions'
    ws['A3'].font = Font(bold=True, size=12)
    ws['A4'] = '12 rue des Innovateurs'
    ws['A5'] = '19100 Brive-la-Gaillarde'
    ws['A6'] = 'Tél: 05 55 17 38 00'
    ws['A7'] = 'Email: contact@techsolutions.fr'
    ws['A8'] = 'SIRET: 801 234 749 00019'
    
    # Informations client
    ws['D3'] = 'CLIENT:'
    ws['D3'].font = Font(bold=True)
    ws['D4'] = 'Raison sociale:'
    ws['E4'] = '[À REMPLIR]'
    ws['D5'] = 'Adresse:'
    ws['E5'] = '[À REMPLIR]'
    ws['D6'] = 'Code postal / Ville:'
    ws['E6'] = '[À REMPLIR]'
    
    # Numéro et date devis
    ws['D8'] = 'N° Devis:'
    ws['E8'] = 'DEVIS-2025-001'
    ws['E8'].font = Font(bold=True)
    ws['D9'] = 'Date:'
    ws['E9'] = datetime.now().strftime('%d/%m/%Y')
    
    # En-têtes tableau
    row = 12
    headers = ['Réf', 'Désignation', 'Qté', 'Prix Unit. HT', 'Montant HT']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row, col, header)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # Lignes de produits (25 lignes vides)
    for i in range(25):
        row_num = 13 + i
        for col in range(1, 6):
            cell = ws.cell(row_num, col)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            # Formules pour les montants
            if col == 5 and i < 25:  # Colonne Montant HT
                cell.value = f'=IF(C{row_num}="","",C{row_num}*D{row_num})'
                cell.number_format = '#,##0.00 €'
    
    # Totaux
    total_row = 39
    ws[f'D{total_row}'] = 'TOTAL HT:'
    ws[f'D{total_row}'].font = Font(bold=True)
    ws[f'D{total_row}'].alignment = Alignment(horizontal='right')
    ws[f'E{total_row}'] = f'=SUM(E13:E37)'
    ws[f'E{total_row}'].font = Font(bold=True)
    ws[f'E{total_row}'].number_format = '#,##0.00 €'
    
    ws[f'D{total_row + 1}'] = 'TVA (20%):'
    ws[f'D{total_row + 1}'].alignment = Alignment(horizontal='right')
    ws[f'E{total_row + 1}'] = f'=E{total_row}*0.2'
    ws[f'E{total_row + 1}'].number_format = '#,##0.00 €'
    
    ws[f'D{total_row + 2}'] = 'TOTAL TTC:'
    ws[f'D{total_row + 2}'].font = Font(bold=True, size=14)
    ws[f'D{total_row + 2}'].fill = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")
    ws[f'D{total_row + 2}'].alignment = Alignment(horizontal='right')
    ws[f'E{total_row + 2}'] = f'=E{total_row}+E{total_row + 1}'
    ws[f'E{total_row + 2}'].font = Font(bold=True, size=14, color="FFFFFF")
    ws[f'E{total_row + 2}'].fill = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")
    ws[f'E{total_row + 2}'].number_format = '#,##0.00 €'
    
    # Conditions
    ws[f'A{total_row + 4}'] = 'CONDITIONS:'
    ws[f'A{total_row + 4}'].font = Font(bold=True)
    ws[f'A{total_row + 5}'] = "Devis valable 30 jours. Paiement à 30 jours fin de mois."
    ws[f'A{total_row + 6}'] = "Délai de livraison: 15 jours ouvrés après acceptation du devis."
    ws[f'A{total_row + 7}'] = "TVA non applicable, art. 293 B du CGI."
    
    # Ajuster les largeurs
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    
    wb.save('/home/claude/techsolutions/Devis_Automatique_TechSolutions.xlsx')
    print("✓ Modèle de devis automatique créé avec succès")
    return wb

if __name__ == "__main__":
    print("=== Génération des fichiers Excel TechSolutions ===\n")
    create_product_catalog()
    create_automated_quote_template()
    print("\n✓ Tous les fichiers ont été générés avec succès!")
    print("\nFichiers créés:")
    print("  - Catalogue_produit_TechSolutions.xlsx")
    print("  - Devis_Automatique_TechSolutions.xlsx")
