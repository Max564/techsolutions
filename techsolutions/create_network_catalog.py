from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
wb.remove(wb.active)

# Couleurs
BLUE = "007BFF"
DARK_BLUE = "0056B3"
LIGHT_BLUE = "E7F3FF"
GRAY = "F5F5F5"
GREEN = "28A745"
ORANGE = "FF6B00"
RED = "DC3545"
WHITE = "FFFFFF"

thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# ==============================================
# FEUILLE 1: FIREWALLS
# ==============================================
ws_fw = wb.create_sheet("Firewalls")
ws_fw.column_dimensions['A'].width = 15
ws_fw.column_dimensions['B'].width = 35
ws_fw.column_dimensions['C'].width = 50
ws_fw.column_dimensions['D'].width = 15
ws_fw.column_dimensions['E'].width = 15
ws_fw.column_dimensions['F'].width = 15

# En-tête
headers = ['Référence', 'Produit', 'Caractéristiques', 'Prix HT', 'Prix TTC', 'Stock']
for col, header in enumerate(headers, 1):
    cell = ws_fw.cell(1, col, header)
    cell.font = Font(bold=True, color=WHITE, size=12)
    cell.fill = PatternFill(start_color=RED, end_color=RED, fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

ws_fw.row_dimensions[1].height = 30

# Données
firewalls = [
    ['FW-FORTI-100F', 'FortiGate 100F', 'Débit 10Gbps, IPS 2.4Gbps, VPN SSL 50 users, IPS/IDS, Antivirus, Web filtering', 3500, 4200, 5],
    ['FW-FORTI-60F', 'FortiGate 60F', 'Débit 5Gbps, IPS 900Mbps, VPN SSL 20 users, Protection DDoS', 1800, 2160, 8],
]

for row_idx, data in enumerate(firewalls, 2):
    for col_idx, value in enumerate(data, 1):
        cell = ws_fw.cell(row_idx, col_idx, value)
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        if col_idx >= 4 and col_idx <= 5:
            cell.number_format = '#,##0 €'
            cell.alignment = Alignment(horizontal='right', vertical='center')
        if col_idx == 6:
            cell.alignment = Alignment(horizontal='center', vertical='center')
    ws_fw.row_dimensions[row_idx].height = 40

# ==============================================
# FEUILLE 2: ROUTEURS
# ==============================================
ws_rtr = wb.create_sheet("Routeurs")
for col in ['A', 'B', 'C', 'D', 'E', 'F']:
    ws_rtr.column_dimensions[col].width = ws_fw.column_dimensions[col].width

for col, header in enumerate(headers, 1):
    cell = ws_rtr.cell(1, col, header)
    cell.font = Font(bold=True, color=WHITE, size=12)
    cell.fill = PatternFill(start_color=BLUE, end_color=BLUE, fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

routeurs = [
    ['RTR-CISCO-4331', 'Cisco ISR 4331', 'Routage 300Mbps, 3x GE, Inter-VLAN, DHCP, QoS, OSPF, VPN', 2800, 3360, 4],
    ['RTR-CISCO-4321', 'Cisco ISR 4321', 'Routage 150Mbps, 2x GE, Inter-VLAN, DHCP, VPN IPsec', 1900, 2280, 6],
]

for row_idx, data in enumerate(routeurs, 2):
    for col_idx, value in enumerate(data, 1):
        cell = ws_rtr.cell(row_idx, col_idx, value)
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        if col_idx >= 4 and col_idx <= 5:
            cell.number_format = '#,##0 €'
            cell.alignment = Alignment(horizontal='right', vertical='center')
        if col_idx == 6:
            cell.alignment = Alignment(horizontal='center', vertical='center')
    ws_rtr.row_dimensions[row_idx].height = 40

# ==============================================
# FEUILLE 3: SWITCHES CORE
# ==============================================
ws_swc = wb.create_sheet("Switches Core")
for col in ['A', 'B', 'C', 'D', 'E', 'F']:
    ws_swc.column_dimensions[col].width = ws_fw.column_dimensions[col].width

for col, header in enumerate(headers, 1):
    cell = ws_swc.cell(1, col, header)
    cell.font = Font(bold=True, color=WHITE, size=12)
    cell.fill = PatternFill(start_color=GREEN, end_color=GREEN, fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

switches_core = [
    ['SW-CISCO-9300-48P', 'Catalyst 9300-48P', '48x GE PoE+ (740W) + 4x SFP+ 10G, Layer 3, 160Gbps, Empilable', 8500, 10200, 3],
    ['SW-CISCO-9300-24P', 'Catalyst 9300-24P', '24x GE PoE+ (435W) + 4x SFP+ 10G, Layer 3, 128Gbps', 5800, 6960, 5],
]

for row_idx, data in enumerate(switches_core, 2):
    for col_idx, value in enumerate(data, 1):
        cell = ws_swc.cell(row_idx, col_idx, value)
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        if col_idx >= 4 and col_idx <= 5:
            cell.number_format = '#,##0 €'
            cell.alignment = Alignment(horizontal='right', vertical='center')
        if col_idx == 6:
            cell.alignment = Alignment(horizontal='center', vertical='center')
    ws_swc.row_dimensions[row_idx].height = 40

# ==============================================
# FEUILLE 4: SWITCHES DISTRIBUTION
# ==============================================
ws_swd = wb.create_sheet("Switches Distribution")
for col in ['A', 'B', 'C', 'D', 'E', 'F']:
    ws_swd.column_dimensions[col].width = ws_fw.column_dimensions[col].width

for col, header in enumerate(headers, 1):
    cell = ws_swd.cell(1, col, header)
    cell.font = Font(bold=True, color=WHITE, size=12)
    cell.fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

switches_dist = [
    ['SW-CISCO-SG350-28P', 'SG350-28P', '24x GE PoE+ (195W) + 2 combo + 2 SFP, Layer 3, 56Gbps', 1200, 1440, 12],
    ['SW-CISCO-SG350-10', 'SG350-10', '8x GE + 2 combo, Layer 3, 20Gbps, Compact', 450, 540, 15],
    ['SW-CISCO-SG250-26P', 'SG250-26P', '24x GE PoE+ (180W) + 2 SFP, Layer 2, 52Gbps', 900, 1080, 10],
]

for row_idx, data in enumerate(switches_dist, 2):
    for col_idx, value in enumerate(data, 1):
        cell = ws_swd.cell(row_idx, col_idx, value)
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        if col_idx >= 4 and col_idx <= 5:
            cell.number_format = '#,##0 €'
            cell.alignment = Alignment(horizontal='right', vertical='center')
        if col_idx == 6:
            cell.alignment = Alignment(horizontal='center', vertical='center')
    ws_swd.row_dimensions[row_idx].height = 40

# ==============================================
# FEUILLE 5: WIFI
# ==============================================
ws_wifi = wb.create_sheet("WiFi")
for col in ['A', 'B', 'C', 'D', 'E', 'F']:
    ws_wifi.column_dimensions[col].width = ws_fw.column_dimensions[col].width

for col, header in enumerate(headers, 1):
    cell = ws_wifi.cell(1, col, header)
    cell.font = Font(bold=True, color=WHITE, size=12)
    cell.fill = PatternFill(start_color=ORANGE, end_color=ORANGE, fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

wifi = [
    ['WIFI-CISCO-2802I', 'Aironet 2802i', 'WiFi 6 (802.11ax), Dual-band, MIMO 4x4, 200 clients, PoE', 800, 960, 20],
    ['WIFI-CISCO-1542I', 'Aironet 1542i', 'WiFi 5 (802.11ac), Dual-band, MIMO 2x2, 100 clients, PoE', 500, 600, 25],
]

for row_idx, data in enumerate(wifi, 2):
    for col_idx, value in enumerate(data, 1):
        cell = ws_wifi.cell(row_idx, col_idx, value)
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        if col_idx >= 4 and col_idx <= 5:
            cell.number_format = '#,##0 €'
            cell.alignment = Alignment(horizontal='right', vertical='center')
        if col_idx == 6:
            cell.alignment = Alignment(horizontal='center', vertical='center')
    ws_wifi.row_dimensions[row_idx].height = 40

# ==============================================
# FEUILLE 6: SERVEURS
# ==============================================
ws_srv = wb.create_sheet("Serveurs")
for col in ['A', 'B', 'C', 'D', 'E', 'F']:
    ws_srv.column_dimensions[col].width = ws_fw.column_dimensions[col].width

for col, header in enumerate(headers, 1):
    cell = ws_srv.cell(1, col, header)
    cell.font = Font(bold=True, color=WHITE, size=12)
    cell.fill = PatternFill(start_color="6C757D", end_color="6C757D", fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

serveurs = [
    ['SRV-DELL-R750', 'Dell PowerEdge R750', '2x Xeon, 128-256GB RAM, 4-8x SSD/HDD, RAID, 2x 10GbE, Rack 2U', 6500, 7800, 6],
    ['SRV-DELL-R250', 'Dell PowerEdge R250', 'Xeon E-2300, 16-64GB RAM, 2-4x SSD/HDD, RAID, Rack 1U', 2500, 3000, 10],
    ['SRV-HPE-DL380', 'HPE ProLiant DL380', '2x Xeon, 128-256GB RAM, 4-12x SSD/HDD, iLO 5, Rack 2U', 7200, 8640, 4],
]

for row_idx, data in enumerate(serveurs, 2):
    for col_idx, value in enumerate(data, 1):
        cell = ws_srv.cell(row_idx, col_idx, value)
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        if col_idx >= 4 and col_idx <= 5:
            cell.number_format = '#,##0 €'
            cell.alignment = Alignment(horizontal='right', vertical='center')
        if col_idx == 6:
            cell.alignment = Alignment(horizontal='center', vertical='center')
    ws_srv.row_dimensions[row_idx].height = 40

# ==============================================
# FEUILLE 7: STOCKAGE NAS
# ==============================================
ws_nas = wb.create_sheet("Stockage NAS")
for col in ['A', 'B', 'C', 'D', 'E', 'F']:
    ws_nas.column_dimensions[col].width = ws_fw.column_dimensions[col].width

for col, header in enumerate(headers, 1):
    cell = ws_nas.cell(1, col, header)
    cell.font = Font(bold=True, color=WHITE, size=12)
    cell.fill = PatternFill(start_color="17A2B8", end_color="17A2B8", fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

nas = [
    ['NAS-SYNOLOGY-RS1221', 'Synology RS1221+', '8 baies 3.5", AMD Ryzen, 4GB RAM, 4x GbE, RAID 0/1/5/6/10, Rack 1U', 3200, 3840, 8],
    ['NAS-QNAP-TS-873A', 'QNAP TS-873A', '8 baies 3.5"/2.5", Ryzen, 8GB RAM, 2x 2.5GbE + 2x GbE, Tour', 2900, 3480, 6],
]

for row_idx, data in enumerate(nas, 2):
    for col_idx, value in enumerate(data, 1):
        cell = ws_nas.cell(row_idx, col_idx, value)
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        if col_idx >= 4 and col_idx <= 5:
            cell.number_format = '#,##0 €'
            cell.alignment = Alignment(horizontal='right', vertical='center')
        if col_idx == 6:
            cell.alignment = Alignment(horizontal='center', vertical='center')
    ws_nas.row_dimensions[row_idx].height = 40

# ==============================================
# FEUILLE 8: ONDULEURS
# ==============================================
ws_ups = wb.create_sheet("Onduleurs UPS")
for col in ['A', 'B', 'C', 'D', 'E', 'F']:
    ws_ups.column_dimensions[col].width = ws_fw.column_dimensions[col].width

for col, header in enumerate(headers, 1):
    cell = ws_ups.cell(1, col, header)
    cell.font = Font(bold=True, color=WHITE, size=12)
    cell.fill = PatternFill(start_color="FFC107", end_color="FFC107", fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

ups = [
    ['UPS-APC-3000', 'APC Smart-UPS 3000VA', '3000VA/2700W, Online, AVR, Hot-swap battery, 8x IEC, Rack 2U', 800, 960, 12],
    ['UPS-EATON-1500', 'Eaton 5P 1550VA', '1550VA/1100W, Line-interactive, AVR, LCD, Rack 1U/Tour', 650, 780, 15],
]

for row_idx, data in enumerate(ups, 2):
    for col_idx, value in enumerate(data, 1):
        cell = ws_ups.cell(row_idx, col_idx, value)
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        if col_idx >= 4 and col_idx <= 5:
            cell.number_format = '#,##0 €'
            cell.alignment = Alignment(horizontal='right', vertical='center')
        if col_idx == 6:
            cell.alignment = Alignment(horizontal='center', vertical='center')
    ws_ups.row_dimensions[row_idx].height = 40

# ==============================================
# FEUILLE 9: KITS RÉSEAU
# ==============================================
ws_kits = wb.create_sheet("Kits Réseau")
ws_kits.column_dimensions['A'].width = 35
ws_kits.column_dimensions['B'].width = 50
ws_kits.column_dimensions['C'].width = 15
ws_kits.column_dimensions['D'].width = 15
ws_kits.column_dimensions['E'].width = 15
ws_kits.column_dimensions['F'].width = 15

headers_kits = ['Kit', 'Description', 'Type', 'Postes', 'Prix HT', 'Remise']
for col, header in enumerate(headers_kits, 1):
    cell = ws_kits.cell(1, col, header)
    cell.font = Font(bold=True, color=WHITE, size=12)
    cell.fill = PatternFill(start_color="28A745", end_color="28A745", fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

kits = [
    ['Kit Réseau TPE', 'Firewall 60F + Routeur 4321 + Switch 26P + 2 WiFi + UPS', 'TPE', 10, 8500, '10%'],
    ['Kit Réseau PME', 'Firewall 100F + Routeur 4331 + Core 9300-24P + 2x SG350-28P + 5 WiFi + 3 UPS', 'PME', 50, 22000, '12%'],
    ['Kit Complet 50 postes', 'Infrastructure + 2x R750 + R250 + NAS Synology 8 baies', 'PME', 50, 58000, '15%'],
]

for row_idx, data in enumerate(kits, 2):
    for col_idx, value in enumerate(data, 1):
        cell = ws_kits.cell(row_idx, col_idx, value)
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        if col_idx == 5:
            cell.number_format = '#,##0 €'
            cell.alignment = Alignment(horizontal='right', vertical='center')
        if col_idx in [3, 4, 6]:
            cell.alignment = Alignment(horizontal='center', vertical='center')
    ws_kits.row_dimensions[row_idx].height = 50

# ==============================================
# FEUILLE 10: RÉSUMÉ
# ==============================================
ws_resume = wb.create_sheet("Résumé", 0)
ws_resume.column_dimensions['A'].width = 30
ws_resume.column_dimensions['B'].width = 15
ws_resume.column_dimensions['C'].width = 20

# Titre
ws_resume.merge_cells('A1:C1')
cell = ws_resume['A1']
cell.value = "CATALOGUE ÉQUIPEMENTS RÉSEAU - TECHSOLUTIONS"
cell.font = Font(bold=True, size=16, color=WHITE)
cell.fill = PatternFill(start_color=BLUE, end_color=BLUE, fill_type='solid')
cell.alignment = Alignment(horizontal='center', vertical='center')
ws_resume.row_dimensions[1].height = 35

ws_resume['A3'] = "Catégorie"
ws_resume['B3'] = "Produits"
ws_resume['C3'] = "Valeur Stock (€ HT)"

for col in ['A', 'B', 'C']:
    cell = ws_resume[f'{col}3']
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color=GRAY, end_color=GRAY, fill_type='solid')
    cell.border = thin_border

categories_resume = [
    ['Firewalls', 2, 5300],
    ['Routeurs', 2, 4700],
    ['Switches Core', 2, 14300],
    ['Switches Distribution', 3, 2550],
    ['Points d''Accès WiFi', 2, 1300],
    ['Serveurs', 3, 16200],
    ['Stockage NAS', 2, 6100],
    ['Onduleurs UPS', 2, 1450],
]

row = 4
for cat, nb, val in categories_resume:
    ws_resume.cell(row, 1, cat).border = thin_border
    ws_resume.cell(row, 2, nb).border = thin_border
    ws_resume.cell(row, 2).alignment = Alignment(horizontal='center')
    ws_resume.cell(row, 3, val).border = thin_border
    ws_resume.cell(row, 3).number_format = '#,##0 €'
    ws_resume.cell(row, 3).alignment = Alignment(horizontal='right')
    row += 1

# Total
ws_resume.cell(row, 1, "TOTAL").font = Font(bold=True)
ws_resume.cell(row, 1).border = thin_border
ws_resume.cell(row, 1).fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type='solid')
ws_resume.cell(row, 2, sum(nb for _, nb, _ in categories_resume)).border = thin_border
ws_resume.cell(row, 2).alignment = Alignment(horizontal='center')
ws_resume.cell(row, 2).font = Font(bold=True)
ws_resume.cell(row, 2).fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type='solid')
ws_resume.cell(row, 3, sum(val for _, _, val in categories_resume)).border = thin_border
ws_resume.cell(row, 3).number_format = '#,##0 €'
ws_resume.cell(row, 3).alignment = Alignment(horizontal='right')
ws_resume.cell(row, 3).font = Font(bold=True)
ws_resume.cell(row, 3).fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type='solid')

wb.save('/mnt/user-data/outputs/Catalogue_Equipements_Reseau_TechSolutions.xlsx')
print("Catalogue équipements réseau créé!")
